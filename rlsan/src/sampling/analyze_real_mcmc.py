"""
加载真实的蒙特卡洛仿真的结果然后进行分析, 计算不同样本规模下的失效率估计值和相对半宽度误差, 并将结果存入xlsx文件中
"""
import numpy as np
import openpyxl
import os
import matplotlib.pyplot as plt
import pickle
from scipy.stats import norm


def load_data(file):
    with open(file, 'rb') as f:
        grid_x = pickle.load(f)
    return grid_x


def failure_indicator(y_data, threshold=0.3):
    # 示性函数,1代表碰撞,0代表安全测试用例
    # 阈值与 IS 脚本保持一致（均为 0.3）
    indicator = (y_data > threshold).astype(float)
    return indicator


def calculate_relative_half_width(failure_estimate, n_samples, conf_level=0.80):
    """
    计算相对半宽度 l_r
    对于 Bernoulli 分布: l_r = z_alpha * sqrt((1 - gamma_hat) / (n * gamma_hat))
    """
    if failure_estimate <= 0:
        return np.inf

    # 计算临界值 z_alpha (例如 80% 置信度对应 alpha=0.2, 则寻找 0.9 分位点)
    z_alpha = norm.ppf(1 - (1 - conf_level) / 2)

    # 计算相对半宽度
    l_r = z_alpha * np.sqrt((1 - failure_estimate) / (n_samples * failure_estimate))
    return l_r


def calculate_cmc_error(failure_estimate, n_samples):
    """
    实现附图公式: e_MC = sqrt((1 - gamma) / (N * gamma))
    用于衡量 CMC 采样的相对标准误差
    """
    if failure_estimate <= 0 or n_samples == 0:
        return np.inf

    e_mc = np.sqrt((1 - failure_estimate) / (n_samples * failure_estimate))
    return e_mc


def calculate_relative_error(gt_faulure, failure_estimate, n_samples):
    if failure_estimate <= 0 or n_samples == 0:
        return np.inf

    e_mc = abs((failure_estimate - gt_faulure) )/ gt_faulure
    return e_mc

def calculate_mcmc_sampling(X_test, y_test, conf_level=0.80):
    """逐批次执行蒙特卡洛采样, 将结果存入xlsx"""

    # 修复: X_test 与 y_test 长度可能不同, 取最小值作为有效样本上界
    # 若只用 len(X_test) 迭代, 当 i > len(y_test) 时 y_test[:i] 仍返回全量 y_test,
    # 导致失效率从该点起不再变化（固定在同一个值）。
    n_total = min(len(X_test), len(y_test))
    if len(X_test) != len(y_test):
        print(f"[警告] X_test 长度({len(X_test)}) 与 y_test 长度({len(y_test)}) 不一致, "
              f"取最小值 {n_total} 进行分析。")

    num_of_samples = []  # 记录当前考察的样本数量
    mcmc_estimates = []  # 记录不同样本下mcmc采样估计失效率
    mcmc_errors = []  # 存储相对半宽度误差

    # 依次从samples中逐次选取批次样本进行评估
    for i in range(1000, n_total + 1, 1000):
        # 读取 0 到 i 的数据
        subset_samples = X_test[:i]
        y_subset = y_test[:i]
        test_restuls = failure_indicator(np.array(y_subset))
        failure = np.mean(test_restuls)
        # error = calculate_cmc_error(failure_estimate, i)
        error = calculate_relative_half_width(failure_estimate=failure, n_samples=i, conf_level=conf_level)

        num_of_samples.append(len(subset_samples))
        mcmc_estimates.append(failure)
        mcmc_errors.append(error)

    # 打印最后一次的全部样本下的失效率估计结果：
    print(f"Final Estimate: {mcmc_estimates[-1]:.6f}")
    print(f"Final Relative error: {mcmc_errors[-1]:.4f}")

    # 绘图
    fig, axes = plt.subplots(2, 1, figsize=(8, 12))
    # ---估计值 ---
    axes[0].plot(num_of_samples, mcmc_estimates, '-o', color='C0', label='mcmc estimate')
    axes[0].set_xlabel('Number of samples')
    axes[0].set_ylabel('Estimated failure rate')
    axes[0].set_title('mcmc estimates vs sample size')
    axes[0].grid(True)
    axes[0].legend()

    # --- 相对半宽度 l_r 曲线 ---
    axes[1].plot(num_of_samples, mcmc_errors, '-s', color='C1', markersize=4, label='Relative errors')
    axes[1].axhline(y=0.05, color='red', linestyle='--', label='Target Accuracy')
    axes[1].set_xlabel('Number of samples')
    axes[1].set_ylabel('CMC relative error')
    axes[1].set_ylim(0.05, min(max(mcmc_errors[2:], default=10.0), 1.0))
    axes[1].grid(True)
    axes[1].legend()

    plt.tight_layout()
    plt.show()


    # 将结果写入xlsx文件
    excel_path = os.path.join('../../results/s1exp/SamplingResults.xlsx')
    workbook = openpyxl.load_workbook(excel_path)
    worksheet = workbook.worksheets[0]
    # 写入测试规模
    for idx, sample_size in enumerate(num_of_samples):
        worksheet.cell(row=idx + 2, column=1, value=sample_size)
    # 写入估计概率
    for idx, failure_estimate in enumerate(mcmc_estimates):
        worksheet.cell(row=idx + 2, column=2, value=failure_estimate)
    # 写入估计误差
    for idx, error in enumerate(mcmc_errors):
        worksheet.cell(row=idx + 2, column=3, value=error)

    workbook.save(excel_path)

if __name__ == "__main__":
    np.random.seed(42)

    # 加载测试结果
    file = '../../results/MCMC/scenario01/sampled_parameters.pkl'
    X_data = load_data(file)
    file = '../../results/s1exp/MCMC_simulation_scores.pkl'
    y_data = load_data(file)

    calculate_mcmc_sampling(X_data, y_data, conf_level=0.80)




