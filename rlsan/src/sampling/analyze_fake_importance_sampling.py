"""
使用训练好的代理模型（GP Surrogate）替代真实 CARLA 仿真，
对重要性采样参数进行失效率估计、SNIS 相对半宽度及 ESS 分析。
与 analyze_real_importance_sampling.py 保持完全一致的误差指标，具备可比性。
"""
import sys
import os
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

import numpy as np
import pickle
import os.path as osp
import openpyxl
import matplotlib.pyplot as plt
import torch
import gpytorch
from scipy.stats import norm
from rlsan.src.surrogate.utils import load_surrogate_model


# 代理模型预测
def surrogate_predict(model, likelihood, X_data, batch_size=5000):
    """分批调用代理模型，返回均值预测分数。"""
    device = next(model.parameters()).device
    n = len(X_data)
    mean_preds = np.zeros(n)

    model.eval()
    likelihood.eval()

    with torch.no_grad(), gpytorch.settings.fast_pred_var():
        for i in range(0, n, batch_size):
            batch_slice = slice(i, min(i + batch_size, n))
            X_batch = torch.tensor(X_data[batch_slice], dtype=torch.float32).to(device)
            posterior = likelihood(model(X_batch))
            mean_preds[batch_slice] = posterior.mean.cpu().numpy()

    return mean_preds


# 工具函数
def get_log_p(x):
    """目标分布（[-1,1]^d 上的均匀分布）的对数密度"""
    dim = x.shape[1]
    vol = 2.0 ** dim
    return np.full(x.shape[0], -np.log(vol))


def failure_indicator(y_data, threshold=0.3):
    """与 CMC 脚本保持一致的阈值"""
    return (y_data > threshold).astype(float)


def calculate_relative_half_width(weights, indicator, failure_estimate, conf_level=0.80):
    """
    SNIS 的 delta-method 相对半宽度：
        l_r = z_α × sqrt(Σwᵢ²(Iᵢ-γ̂)²) / (Σwᵢ × γ̂)
    与 analyze_real_importance_sampling.py 完全相同。
    """
    if failure_estimate <= 0 or len(weights) < 2:
        return np.inf
    z_alpha = norm.ppf(1 - (1 - conf_level) / 2)
    weighted_diff = weights * (indicator - failure_estimate)
    standard_error = np.sqrt(np.sum(weighted_diff ** 2)) / np.sum(weights)
    return (z_alpha * standard_error) / failure_estimate


# 主分析函数
def calculate_fake_importance_sampling(X_data, y_surrogate, q_mix_values, conf_level=0.80):
    """
    基于代理模型预测结果执行 SNIS，逐步计算失效率、相对半宽度和 ESS/N。
    与 analyze_real_importance_sampling.py 的分析流程完全对齐。
    """
    total_N = len(X_data)
    log_p = get_log_p(X_data)
    log_q = np.log(q_mix_values + 1e-20)
    log_weights_all = log_p - log_q
    indicator_all = failure_indicator(y_surrogate)

    num_of_samples = []
    IS_estimates = []
    IS_l_r_history = []
    IS_ess_ratios = []

    print(f"\n--- Surrogate-based IS Analysis ({conf_level*100:.0f}% Confidence) ---")

    step = 100
    for N in range(step, total_N + 1, step):
        subset_log_w = log_weights_all[:N]
        subset_indicator = indicator_all[:N]

        # 数值稳定的权重
        max_log_w = np.max(subset_log_w)
        weights_shifted = np.exp(subset_log_w - max_log_w)

        # 权重截断（99th percentile）
        clip_threshold = np.quantile(weights_shifted, 0.99)
        weights_clipped = np.clip(weights_shifted, 0, clip_threshold)

        # 1. SNIS 失效率估计
        sum_w = np.sum(weights_clipped)
        failure_estimate = np.sum(weights_clipped * subset_indicator) / sum_w if sum_w > 0 else 0

        # 2. SNIS 相对半宽度（delta method）
        error = calculate_relative_half_width(weights_clipped, subset_indicator,
                                              failure_estimate, conf_level)

        # 3. ESS/N
        ess = np.sum(weights_clipped) ** 2 / np.sum(weights_clipped ** 2)
        ess_ratio = ess / N

        IS_estimates.append(failure_estimate)
        IS_l_r_history.append(error)
        IS_ess_ratios.append(ess_ratio)
        num_of_samples.append(N)

    final_ess_ratio = IS_ess_ratios[-1]
    print(f"Final Estimate:  {IS_estimates[-1]:.6f}")
    print(f"Final l_r:       {IS_l_r_history[-1]:.4f}")
    print(f"Final ESS/N:     {final_ess_ratio:.4f}  "
          f"({'良好' if final_ess_ratio > 0.1 else '偏低，提案分布与目标分布失配严重'})")

    # 绘图
    fig, axes = plt.subplots(2, 1, figsize=(8, 10))

    axes[0].plot(num_of_samples, IS_estimates, color='C0', label='Surrogate IS Estimate')
    axes[0].set_ylabel('Failure Rate')
    axes[0].set_title('Surrogate IS: Failure Rate Convergence')
    axes[0].grid(True, alpha=0.3)
    axes[0].legend()

    axes[1].plot(num_of_samples, IS_l_r_history, color='C1', label='Relative Half-width $l_r$')
    axes[1].axhline(y=0.05, color='red', linestyle='--', label='Target $l_r=0.05$')
    axes[1].set_xlabel('Sample Size')
    axes[1].set_ylabel('$l_r$')
    axes[1].set_title('Surrogate IS: Statistical Precision')
    axes[1].grid(True, alpha=0.3)
    axes[1].legend()

    plt.tight_layout()
    plt.savefig('../../results/s1exp/Surrogate_IS_Analysis.png', dpi=150)
    plt.show()

    # Excel 写入（Sheet 2，与 real IS 同 Sheet）
    excel_path = '../../results/s1exp/SamplingResults.xlsx'
    try:
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook.worksheets[1]
        for idx in range(len(num_of_samples)):
            worksheet.cell(row=idx + 2, column=1, value=num_of_samples[idx])
            worksheet.cell(row=idx + 2, column=2, value=IS_estimates[idx])
            worksheet.cell(row=idx + 2, column=3, value=IS_l_r_history[idx])
        workbook.save(excel_path)
        print(f"Results saved to {excel_path} (Sheet 2)")
    except Exception as e:
        print(f"Excel save failed: {e}")


# 入口
if __name__ == "__main__":
    np.random.seed(42)

    # 1. 加载代理模型
    model_path = '../../results/s1exp/updated_surrogate_model.pkl'
    model, likelihood = load_surrogate_model(model_path)

    # 2. 加载 IS 采样参数与提案密度包
    params_file = osp.join('../../results/IS/scenario01', 'sampled_parameters.pkl')
    package_file = '../../results/s1exp/final_refined_IS_package.pkl'

    with open(params_file, 'rb') as f:
        X_data = pickle.load(f)
    with open(package_file, 'rb') as f:
        package = pickle.load(f)

    log_q_mix_values = package['log_q_mix_values']
    q_mix_values = np.exp(log_q_mix_values)

    # 3. 长度对齐
    min_len = min(len(X_data), len(q_mix_values))
    X_data = X_data[:min_len]
    q_mix_values = q_mix_values[:min_len]
    print(f"Loaded IS samples: {X_data.shape}")

    # 4. 代理模型预测
    print("Running surrogate prediction...")
    y_surrogate = surrogate_predict(model, likelihood, X_data)
    print(f"Prediction done. Score range: [{y_surrogate.min():.4f}, {y_surrogate.max():.4f}]")

    # 5. 分析
    calculate_fake_importance_sampling(X_data, y_surrogate, q_mix_values, conf_level=0.80)
