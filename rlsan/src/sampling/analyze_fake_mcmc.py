"""
使用训练好的代理模型（GP Surrogate）替代真实 CARLA 仿真，
对蒙特卡洛采样参数进行失效率估计和误差分析。
结果与 analyze_real_mcmc.py 保持一致的误差指标，具备可比性。
"""
import sys
import os
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__)), '..'))

import numpy as np
import pickle
import openpyxl
import matplotlib.pyplot as plt
import torch
import gpytorch
from scipy.stats import norm
from rlsan.src.surrogate.utils import load_surrogate_model


# 代理模型预测
def surrogate_predict(model, likelihood, X_data, batch_size=5000):
    """
    使用代理模型对全量样本分批预测，返回均值预测分数。
    与 verify_surrogate.py 中的 global_predict 逻辑一致。
    """
    device = next(model.parameters()).device
    n = len(X_data)
    mean_preds = np.zeros(n)

    model.eval()
    likelihood.eval()

    with torch.no_grad(), gpytorch.settings.fast_pred_var():
        for i in range(0, n, batch_size):
            batch_slice = slice(i, min(i + batch_size, n))
            X_batch = torch.tensor(X_data[batch_slice], dtype=torch.float32).to(device)
            posterior = likelihood(model(X_batch))
            mean_preds[batch_slice] = posterior.mean.cpu().numpy()

    return mean_preds


# 失效判别与误差公式
def failure_indicator(y_data, threshold=0.3):
    """示性函数，阈值与 IS 脚本保持一致（均为 0.3）"""
    return (y_data > threshold).astype(float)


def calculate_relative_half_width(failure_estimate, n_samples, conf_level=0.80):
    """
    Bernoulli 分布的相对半宽度：
        l_r = z_α × sqrt((1 - γ̂) / (N × γ̂))
    与 IS 脚本的 delta-method 在权重全为 1 时数学等价。
    """
    if failure_estimate <= 0:
        return np.inf
    z_alpha = norm.ppf(1 - (1 - conf_level) / 2)
    return z_alpha * np.sqrt((1 - failure_estimate) / (n_samples * failure_estimate))


# 主分析函数
def calculate_fake_mcmc_sampling(X_test, y_surrogate, conf_level=0.80):
    """
    逐批次分析代理模型预测结果，计算失效率估计和相对半宽度，
    输出结果与 analyze_real_mcmc.py 完全对齐，具备横向可比性。
    """
    n_total = len(X_test)
    indicator_all = failure_indicator(y_surrogate)

    num_of_samples = []
    estimates = []
    errors = []

    for i in range(1000, n_total + 1, 1000):
        subset_indicator = indicator_all[:i]
        failure = np.mean(subset_indicator)
        error = calculate_relative_half_width(failure, i, conf_level)

        num_of_samples.append(i)
        estimates.append(failure)
        errors.append(error)

    print(f"\n--- Surrogate-based CMC Analysis ({conf_level*100:.0f}% Confidence) ---")
    print(f"Final Estimate: {estimates[-1]:.6f}")
    print(f"Final l_r:      {errors[-1]:.4f}")

    # 绘图
    fig, axes = plt.subplots(2, 1, figsize=(8, 12))

    axes[0].plot(num_of_samples, estimates, '-o', color='C0', markersize=3,
                 label='Surrogate CMC Estimate')
    axes[0].set_xlabel('Number of Samples')
    axes[0].set_ylabel('Estimated Failure Rate')
    axes[0].set_title('Surrogate CMC: Failure Rate vs Sample Size')
    axes[0].grid(True)
    axes[0].legend()

    axes[1].plot(num_of_samples, errors, '-s', color='C1', markersize=3,
                 label='Relative Half-width $l_r$')
    axes[1].axhline(y=0.05, color='red', linestyle='--', label='Target $l_r=0.05$')
    axes[1].set_xlabel('Number of Samples')
    axes[1].set_ylabel('$l_r$')
    axes[1].set_title('Surrogate CMC: Statistical Precision')
    axes[1].grid(True)
    axes[1].legend()

    plt.tight_layout()
    plt.savefig('../../results/s1exp/Surrogate_CMC_Analysis.png', dpi=150)
    plt.show()

    # Excel 写入（Sheet 1，与 real_mcmc 同 Sheet）
    excel_path = '../../results/s1exp/SamplingResults.xlsx'
    try:
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook.worksheets[0]
        for idx in range(len(num_of_samples)):
            worksheet.cell(row=idx + 2, column=1, value=num_of_samples[idx])
            worksheet.cell(row=idx + 2, column=2, value=estimates[idx])
            worksheet.cell(row=idx + 2, column=3, value=errors[idx])
        workbook.save(excel_path)
        print(f"Results saved to {excel_path} (Sheet 1)")
    except Exception as e:
        print(f"Excel save failed: {e}")


if __name__ == "__main__":
    np.random.seed(42)

    # 1. 加载代理模型
    model_path = '../../results/s1exp/updated_surrogate_model.pkl'
    model, likelihood = load_surrogate_model(model_path)

    # 2. 加载蒙特卡洛采样参数
    params_file = '../../results/MCMC/scenario01/sampled_parameters.pkl'
    with open(params_file, 'rb') as f:
        X_data = pickle.load(f)
    print(f"Loaded MCMC samples: {X_data.shape}")

    # 3. 代理模型预测
    print("Running surrogate prediction...")
    y_surrogate = surrogate_predict(model, likelihood, X_data)
    print(f"Prediction done. Score range: [{y_surrogate.min():.4f}, {y_surrogate.max():.4f}]")

    # 4. 分析
    calculate_fake_mcmc_sampling(X_data, y_surrogate, conf_level=0.80)
