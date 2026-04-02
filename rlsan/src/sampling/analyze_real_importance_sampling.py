"""
读取实际的重要性采样的仿真实验结果,评价加速性和估计结果的无偏性
集成相对半宽度 (Relative Half-width) 作为误差估计
"""
import numpy as np
import pickle
import os
import os.path as osp
import openpyxl
import matplotlib.pyplot as plt
from scipy.stats import norm
# 使用 logsumexp 保证数值稳定
from scipy.special import logsumexp


def load_data(file):
    with open(file, 'rb') as f:
        data = pickle.load(f)
    return data


def get_log_p(x):
    """计算目标分布（均匀分布）的对数概率"""
    dim = x.shape[1]
    vol = 2.0 ** dim
    return np.full(x.shape[0], -np.log(vol))


def failure_indicator(y_data, threshold=0.3):
    return (y_data > threshold).astype(float)


def calculate_relative_half_width(weights, indicator, failure_estimate, conf_level=0.80):
    """
    计算重要性采样的相对半宽度 l_r
    针对 SNIS 的方差估算：l_r = z_alpha * std(weighted_samples) / (mean * sqrt(n))
    """
    if failure_estimate <= 0 or len(weights) < 2:
        return np.inf

    z_alpha = norm.ppf(1 - (1 - conf_level) / 2)

    # 针对 SNIS 的方差估算 (Delta Method 简化版)
    # v_i = w_i * (I_i - gamma_hat)
    weighted_diff = weights * (indicator - failure_estimate)

    # 估计值的标准差
    # s_error = sqrt( sum(v_i^2) ) / sum(w_i)
    standard_error = np.sqrt(np.sum(weighted_diff**2)) / np.sum(weights)

    # 相对半宽度 = (z_alpha * 标准差) / 估计值
    l_r = (z_alpha * standard_error) / failure_estimate
    return l_r


def calculate_IS_error(weights, indicator, failure_estimate):
    """
    e_IS = (1/N) * sqrt( sum( ( (I*f)/(E*h) - 1 )^2 ) )
    注意：f/h 就是权重 weights，E 是 true_value
    """
    N = len(weights)
    if N == 0 or failure_estimate == 0:
        return np.inf
    
    # 公式内部项: (I(x) * w(x) / E) - 1
    # 这里的 weights = f(x)/h(x)
    term = (indicator * weights / failure_estimate) - 1
    
    # 计算平方和的开方，再除以 N
    e_is = (1.0 / N) * np.sqrt(np.sum(term**2))
    
    return e_is


def calculate_importance_sampling(X_data, y_data, q_mix_values, conf_level=0.80):
    """使用数值稳定的 SNIS 评估失效率及相对半宽度"""

    num_of_samples = []
    IS_estimates = []
    IS_l_r_history = []  # 存储相对半宽度
    IS_ess_ratios = []   # 存储 ESS/N 比例

    total_N = len(X_data)
    log_p = get_log_p(X_data)
    log_q = np.log(q_mix_values + 1e-20)
    log_weights_all = log_p - log_q
    indicator_all = failure_indicator(np.array(y_data))

    print(f"\n--- Starting IS Precision Analysis ({conf_level*100}% Confidence) ---")

    step = 100
    for N in range(step, total_N + 1, step):
        subset_log_w = log_weights_all[:N]
        subset_indicator = indicator_all[:N]

        # 数值稳定的权重计算
        max_log_w = np.max(subset_log_w)
        weights_shifted = np.exp(subset_log_w - max_log_w)

        # 权重截断：将极端权重截断到第 99 百分位，以小偏差换取大幅降低方差
        # 这是 IS 领域的标准做法 (Ionides 2008, Owen 2013)
        clip_threshold = np.quantile(weights_shifted, 0.99)
        weights_clipped = np.clip(weights_shifted, 0, clip_threshold)

        # 1. 估计失效率 (SNIS with clipped weights)
        sum_w = np.sum(weights_clipped)
        failure_estimate = np.sum(weights_clipped * subset_indicator) / sum_w if sum_w > 0 else 0

        # 2. 计算 SNIS 相对半宽度 (delta method, 比 calculate_IS_error 更稳定)
        error = calculate_relative_half_width(
            weights=weights_clipped,
            indicator=subset_indicator,
            failure_estimate=failure_estimate,
            conf_level=conf_level,
        )

        # 3. 计算有效样本量 ESS 及其占比
        # ESS = (Σwᵢ)² / Σwᵢ²  反映权重分布的均匀程度，ESS/N 越接近 1 越好
        ess = np.sum(weights_clipped) ** 2 / np.sum(weights_clipped ** 2)
        ess_ratio = ess / N

        IS_estimates.append(failure_estimate)
        IS_l_r_history.append(error)
        IS_ess_ratios.append(ess_ratio)
        num_of_samples.append(N)

    final_ess_ratio = IS_ess_ratios[-1]
    print(f"Final Estimate:  {IS_estimates[-1]:.6f}")
    print(f"Final l_r:       {IS_l_r_history[-1]:.4f}")
    print(f"Final ESS/N:     {final_ess_ratio:.4f}  "
          f"({'良好' if final_ess_ratio > 0.1 else '偏低，提案分布与目标分布失配严重'})")

    # --- 绘图 ---
    fig, axes = plt.subplots(2, 1, figsize=(8, 10))
    axes[0].plot(num_of_samples, IS_estimates, color='C0', label='IS Estimate')
    axes[0].set_ylabel('Failure Rate')
    axes[0].set_title('IS Failure Rate Convergence')
    axes[0].grid(True, alpha=0.3)
    axes[0].legend()

    axes[1].plot(num_of_samples, IS_l_r_history, color='C1', label='Relative error')
    axes[1].set_xlabel('Sample Size')
    axes[1].set_ylabel('l_r ')
    axes[1].set_title(f'Statistical Precision ($l_r$)')
    axes[1].grid(True, alpha=0.3)
    axes[1].legend()

    plt.tight_layout()
    plt.savefig('../../results/s1exp/IS_Precision_Analysis.png')
    plt.show()

    # --- Excel 写入 ---
    excel_path = '../../results/s1exp/SamplingResults.xlsx'
    try:
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook.worksheets[1]
        for idx in range(len(num_of_samples)):
            worksheet.cell(row=idx + 2, column=1, value=num_of_samples[idx])
            worksheet.cell(row=idx + 2, column=2, value=IS_estimates[idx])
            worksheet.cell(row=idx + 2, column=3, value=IS_l_r_history[idx])
        workbook.save(excel_path)
    except Exception as e:
        print(f"Excel save failed: {e}")


if __name__ == "__main__":
    np.random.seed(42)

    # 路径根据实际情况调整
    params_file = osp.join('../../results/IS/scenario01', 'sampled_parameters.pkl')
    scores_file = osp.join('../../results/s1exp', 'IS_simulation_scores.pkl')
    package = load_data( '../../results/s1exp/final_refined_IS_package.pkl')

    X_data = load_data(params_file)
    y_data = load_data(scores_file)
    log_q_mix_values = package['log_q_mix_values']
    q_mix_values = np.exp(log_q_mix_values)

    min_len = min(len(X_data), len(y_data), len(q_mix_values))
    calculate_importance_sampling(X_data[:min_len], y_data[:min_len], q_mix_values[:min_len], conf_level=0.80)